import json
import csv
from pathlib import Path

# ============================================================
# CONFIGURAÇÃO
# ============================================================

PASTA_SCRIPT = Path(__file__).resolve().parent
RAIZ_PROJETO = PASTA_SCRIPT.parents[2]

FICHEIRO_FRAGMENTOS = (
    RAIZ_PROJETO
    / "13_Meta_Indice"
    / "cadência"
    / "01_segmentar_fragmentos"
    / "fragmentos_resegmentados.json"
)

FICHEIRO_CADENCIA = (
    RAIZ_PROJETO
    / "13_Meta_Indice"
    / "cadência"
    / "02_extrator_cadência"
    / "cadencia_extraida.json"
)

CSV_SAIDA = PASTA_SCRIPT / "tabela_revisao_cadencia.csv"
XLSX_SAIDA = PASTA_SCRIPT / "tabela_revisao_cadencia.xlsx"

CSV_REVISAO_HUMANA = PASTA_SCRIPT / "revisao_humana_prioritaria.csv"
CSV_CONFIANCA_BAIXA = PASTA_SCRIPT / "confianca_baixa.csv"
CSV_ZONA_INDEFINIDA = PASTA_SCRIPT / "zona_indefinida.csv"
CSV_MATERIAL_RETRABALHAR = PASTA_SCRIPT / "material_a_retrabalhar.csv"

def validar_ficheiro(caminho: Path, nome: str):
    if not caminho.exists():
        raise FileNotFoundError(f"{nome} não encontrado: {caminho}")


validar_ficheiro(FICHEIRO_FRAGMENTOS, "FICHEIRO_FRAGMENTOS")
validar_ficheiro(FICHEIRO_CADENCIA, "FICHEIRO_CADENCIA")

# ============================================================
# UTILITÁRIOS
# ============================================================

def carregar_json(caminho: Path):
    with caminho.open("r", encoding="utf-8") as f:
        return json.load(f)


def lista_para_texto(valor):
    if valor is None:
        return ""
    if isinstance(valor, list):
        return " | ".join(str(x) for x in valor)
    return str(valor)


def valor_ou_vazio(d, chave, default=""):
    if not isinstance(d, dict):
        return default
    valor = d.get(chave, default)
    return "" if valor is None else valor


def prioridade_confianca(valor):
    """
    Para ordenar:
    baixa -> 0
    media -> 1
    alta  -> 2
    resto -> 9
    """
    mapa = {
        "baixa": 0,
        "media": 1,
        "alta": 2,
    }
    return mapa.get(str(valor).strip().lower(), 9)


def prioridade_zona(valor):
    """
    Para ordenar:
    indefinida primeiro.
    """
    return 0 if str(valor).strip().lower() == "indefinida" else 1


def prioridade_aproveitamento(valor):
    """
    material_a_retrabalhar primeiro.
    """
    return 0 if str(valor).strip().lower() == "material_a_retrabalhar" else 1


def prioridade_bool_true_primeiro(valor):
    return 0 if str(valor).strip().lower() == "true" else 1


def escrever_csv(caminho: Path, linhas, colunas):
    with caminho.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=colunas, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(linhas)


def tentar_escrever_xlsx(caminho: Path, linhas, colunas):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl não está disponível. Ficheiro XLSX não será criado.")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "RevisaoCadencia"

    # Cabeçalhos
    ws.append(colunas)

    # Linhas
    for linha in linhas:
        ws.append([linha.get(col, "") for col in colunas])

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Auto-largura simples
    for i, col in enumerate(colunas, start=1):
        max_len = len(col)
        for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            valor = row[0].value
            if valor is None:
                continue
            max_len = max(max_len, len(str(valor)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

    wb.save(caminho)
    return True


# ============================================================
# CARGA DOS DADOS
# ============================================================

fragmentos = carregar_json(FICHEIRO_FRAGMENTOS)
cadencias = carregar_json(FICHEIRO_CADENCIA)

if not isinstance(fragmentos, list):
    raise ValueError("fragmentos_resegmentados.json não tem uma lista no topo.")

if not isinstance(cadencias, list):
    raise ValueError("cadencia_extraida.json não tem uma lista no topo.")


# ============================================================
# INDEXAÇÃO
# ============================================================

fragmentos_por_id = {}
for frag in fragmentos:
    fragment_id = frag.get("fragment_id")
    if fragment_id:
        fragmentos_por_id[fragment_id] = frag

cadencias_por_id = {}
for cad in cadencias:
    fragment_id = cad.get("fragment_id")
    if fragment_id:
        cadencias_por_id[fragment_id] = cad


# ============================================================
# CRUZAMENTO
# ============================================================

todos_ids = sorted(set(fragmentos_por_id.keys()) | set(cadencias_por_id.keys()))

linhas = []

for fragment_id in todos_ids:
    frag = fragmentos_por_id.get(fragment_id, {})
    cad = cadencias_por_id.get(fragment_id, {})

    origem = frag.get("origem", {}) if isinstance(frag, dict) else {}
    segmentacao = frag.get("segmentacao", {}) if isinstance(frag, dict) else {}
    integridade = frag.get("integridade_semantica", {}) if isinstance(frag, dict) else {}
    relacoes = frag.get("relacoes_locais", {}) if isinstance(frag, dict) else {}
    sinalizador = frag.get("sinalizador_para_cadencia", {}) if isinstance(frag, dict) else {}
    meta_seg = frag.get("_metadados_segmentador", {}) if isinstance(frag, dict) else {}

    bloco_cad = cad.get("cadencia", {}) if isinstance(cad, dict) else {}

    linha = {
        # Identificação
        "fragment_id": fragment_id,
        "origem_id": cad.get("origem_id") or origem.get("origem_id", ""),
        "ordem_no_ficheiro": cad.get("ordem_no_ficheiro") or origem.get("ordem_no_ficheiro", ""),
        "ficheiro_origem": origem.get("ficheiro", ""),
        "data_origem": origem.get("data", ""),
        "header_original": origem.get("header_original", ""),

        # Texto
        "texto_fragmento": frag.get("texto_fragmento", ""),
        "texto_normalizado": frag.get("texto_normalizado", ""),
        "n_chars_fragmento": frag.get("n_chars_fragmento", ""),
        "paragrafos_agregados": frag.get("paragrafos_agregados", ""),
        "frases_aproximadas": frag.get("frases_aproximadas", ""),
        "densidade_aprox": frag.get("densidade_aprox", ""),

        # Segmentação
        "tipo_material_fonte": frag.get("tipo_material_fonte", ""),
        "tipo_unidade_segmentacao": valor_ou_vazio(segmentacao, "tipo_unidade"),
        "criterio_de_unidade": valor_ou_vazio(segmentacao, "criterio_de_unidade"),
        "houve_fusao_de_paragrafos": valor_ou_vazio(segmentacao, "houve_fusao_de_paragrafos"),
        "houve_corte_interno": valor_ou_vazio(segmentacao, "houve_corte_interno"),
        "container_tipo_segmentacao": valor_ou_vazio(segmentacao, "container_tipo_segmentacao"),

        "funcao_textual_dominante": frag.get("funcao_textual_dominante", ""),
        "tema_dominante_provisorio": frag.get("tema_dominante_provisorio", ""),
        "conceitos_relevantes_provisorios": lista_para_texto(frag.get("conceitos_relevantes_provisorios")),
        "integridade_semantica_grau": valor_ou_vazio(integridade, "grau"),
        "confianca_segmentacao": frag.get("confianca_segmentacao", ""),

        "fragmento_anterior": valor_ou_vazio(relacoes, "fragmento_anterior"),
        "fragmento_seguinte": valor_ou_vazio(relacoes, "fragmento_seguinte"),
        "continua_anterior": valor_ou_vazio(relacoes, "continua_anterior"),
        "prepara_seguinte_segmentacao": valor_ou_vazio(relacoes, "prepara_seguinte"),

        "estado_revisao_segmentacao": frag.get("estado_revisao", ""),
        "pronto_para_extrator_cadencia": valor_ou_vazio(sinalizador, "pronto_para_extrator_cadencia"),
        "requer_revisao_manual_prioritaria_segmentacao": valor_ou_vazio(sinalizador, "requer_revisao_manual_prioritaria"),

        "modelo_segmentador": valor_ou_vazio(meta_seg, "modelo"),
        "versao_segmentador": valor_ou_vazio(meta_seg, "versao_segmentador"),
        "timestamp_segmentador": valor_ou_vazio(meta_seg, "timestamp"),

        # Cadência
        "funcao_cadencia_principal": valor_ou_vazio(bloco_cad, "funcao_cadencia_principal"),
        "funcao_cadencia_secundaria": valor_ou_vazio(bloco_cad, "funcao_cadencia_secundaria"),
        "direcao_movimento": valor_ou_vazio(bloco_cad, "direcao_movimento"),
        "grau_de_abertura_argumentativa": valor_ou_vazio(bloco_cad, "grau_de_abertura_argumentativa"),
        "centralidade": valor_ou_vazio(bloco_cad, "centralidade"),
        "estatuto_no_percurso": valor_ou_vazio(bloco_cad, "estatuto_no_percurso"),
        "zona_provavel_percurso": valor_ou_vazio(bloco_cad, "zona_provavel_percurso"),
        "ponte_entre_zonas": valor_ou_vazio(bloco_cad, "ponte_entre_zonas"),
        "prepara_fragmento_seguinte_cadencia": valor_ou_vazio(bloco_cad, "prepara_fragmento_seguinte"),
        "fecha_algo_anterior": valor_ou_vazio(bloco_cad, "fecha_algo_anterior"),
        "incide_sobre_erro": valor_ou_vazio(bloco_cad, "incide_sobre_erro"),
        "familia_erro_provavel": valor_ou_vazio(bloco_cad, "familia_erro_provavel"),
        "erro_dominante_provavel": valor_ou_vazio(bloco_cad, "erro_dominante_provavel"),
        "dominio_contributivo_principal": valor_ou_vazio(bloco_cad, "dominio_contributivo_principal"),
        "tipo_fragmento_provavel": valor_ou_vazio(bloco_cad, "tipo_fragmento_provavel"),
        "aproveitamento_editorial": valor_ou_vazio(bloco_cad, "aproveitamento_editorial"),
        "necessita_revisao_humana": valor_ou_vazio(bloco_cad, "necessita_revisao_humana"),
        "confianca_cadencia": valor_ou_vazio(bloco_cad, "confianca_cadencia"),
        "justificacao_curta_cadencia": valor_ou_vazio(bloco_cad, "justificacao_curta_cadencia"),

        # Auxiliares de ordenação
        "_ord_revisao_humana": prioridade_bool_true_primeiro(valor_ou_vazio(bloco_cad, "necessita_revisao_humana")),
        "_ord_confianca": prioridade_confianca(valor_ou_vazio(bloco_cad, "confianca_cadencia")),
        "_ord_zona": prioridade_zona(valor_ou_vazio(bloco_cad, "zona_provavel_percurso")),
        "_ord_aproveitamento": prioridade_aproveitamento(valor_ou_vazio(bloco_cad, "aproveitamento_editorial")),
    }

    linhas.append(linha)


# ============================================================
# ORDENAÇÃO
# ============================================================

def chave_ordenacao(linha):
    ordem = linha.get("ordem_no_ficheiro", "")
    try:
        ordem_num = int(ordem)
    except Exception:
        ordem_num = 999999

    return (
        linha["_ord_revisao_humana"],
        linha["_ord_confianca"],
        linha["_ord_zona"],
        linha["_ord_aproveitamento"],
        ordem_num,
        linha.get("fragment_id", ""),
    )


linhas.sort(key=chave_ordenacao)


# ============================================================
# REMOVER COLUNAS AUXILIARES DA SAÍDA
# ============================================================

colunas = [
    "fragment_id",
    "origem_id",
    "ordem_no_ficheiro",
    "ficheiro_origem",
    "data_origem",
    "header_original",

    "texto_fragmento",
    "texto_normalizado",
    "n_chars_fragmento",
    "paragrafos_agregados",
    "frases_aproximadas",
    "densidade_aprox",

    "tipo_material_fonte",
    "tipo_unidade_segmentacao",
    "criterio_de_unidade",
    "houve_fusao_de_paragrafos",
    "houve_corte_interno",
    "container_tipo_segmentacao",
    "funcao_textual_dominante",
    "tema_dominante_provisorio",
    "conceitos_relevantes_provisorios",
    "integridade_semantica_grau",
    "confianca_segmentacao",
    "fragmento_anterior",
    "fragmento_seguinte",
    "continua_anterior",
    "prepara_seguinte_segmentacao",
    "estado_revisao_segmentacao",
    "pronto_para_extrator_cadencia",
    "requer_revisao_manual_prioritaria_segmentacao",
    "modelo_segmentador",
    "versao_segmentador",
    "timestamp_segmentador",

    "funcao_cadencia_principal",
    "funcao_cadencia_secundaria",
    "direcao_movimento",
    "grau_de_abertura_argumentativa",
    "centralidade",
    "estatuto_no_percurso",
    "zona_provavel_percurso",
    "ponte_entre_zonas",
    "prepara_fragmento_seguinte_cadencia",
    "fecha_algo_anterior",
    "incide_sobre_erro",
    "familia_erro_provavel",
    "erro_dominante_provavel",
    "dominio_contributivo_principal",
    "tipo_fragmento_provavel",
    "aproveitamento_editorial",
    "necessita_revisao_humana",
    "confianca_cadencia",
    "justificacao_curta_cadencia",
]

linhas_limpa = [{col: linha.get(col, "") for col in colunas} for linha in linhas]


# ============================================================
# ESCREVER SAÍDAS PRINCIPAIS
# ============================================================

escrever_csv(CSV_SAIDA, linhas_limpa, colunas)
xlsx_ok = tentar_escrever_xlsx(XLSX_SAIDA, linhas_limpa, colunas)

print(f"CSV principal criado: {CSV_SAIDA}")
if xlsx_ok:
    print(f"XLSX principal criado: {XLSX_SAIDA}")


# ============================================================
# ESCREVER SUBCONJUNTOS PRIORITÁRIOS
# ============================================================

revisao_humana = [
    l for l in linhas_limpa
    if str(l.get("necessita_revisao_humana", "")).strip().lower() == "true"
]

confianca_baixa = [
    l for l in linhas_limpa
    if str(l.get("confianca_cadencia", "")).strip().lower() == "baixa"
]

zona_indefinida = [
    l for l in linhas_limpa
    if str(l.get("zona_provavel_percurso", "")).strip().lower() == "indefinida"
]

material_a_retrabalhar = [
    l for l in linhas_limpa
    if str(l.get("aproveitamento_editorial", "")).strip().lower() == "material_a_retrabalhar"
]

escrever_csv(CSV_REVISAO_HUMANA, revisao_humana, colunas)
escrever_csv(CSV_CONFIANCA_BAIXA, confianca_baixa, colunas)
escrever_csv(CSV_ZONA_INDEFINIDA, zona_indefinida, colunas)
escrever_csv(CSV_MATERIAL_RETRABALHAR, material_a_retrabalhar, colunas)

print(f"CSV revisão humana: {CSV_REVISAO_HUMANA} ({len(revisao_humana)} linhas)")
print(f"CSV confiança baixa: {CSV_CONFIANCA_BAIXA} ({len(confianca_baixa)} linhas)")
print(f"CSV zona indefinida: {CSV_ZONA_INDEFINIDA} ({len(zona_indefinida)} linhas)")
print(f"CSV material a retrabalhar: {CSV_MATERIAL_RETRABALHAR} ({len(material_a_retrabalhar)} linhas)")

print("Concluído.")